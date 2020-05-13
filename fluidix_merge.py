import sys
import argparse
import uuid
import openpyxl
import csv
import os
import pandas as pd
import string


def cmdline_args():
    # Make parser object
    p = argparse.ArgumentParser(description="""
        This is a test of the command line argument parser in Python.
        """,
                                formatter_class=argparse.ArgumentDefaultsHelpFormatter)

    # Optional argument which requires a parameter (eg. -d test)
    p.add_argument("-i", "--input_filename", required=True)
    p.add_argument("-t", "--template_filename", required=True)
    p.add_argument("-p", "--plate_type", type=int,
                   choices=[96, 48, 196], default=96, required=True)
    p.add_argument("-op", "--output_prefix", required=True)
    p.add_argument("-od", "--output_directory", required=True)

    return(p.parse_args())


# Try running with these args
#
# "Hello" 123 --enable
if __name__ == '__main__':

    """
    if sys.version_info<(3,0,0):
        sys.stderr.write("You need python 3.0 or later to run this script\n")
        sys.exit(1)
    """
    args = cmdline_args()
    inputfile = args.input_filename
    template = args.template_filename
    plate_type = args.plate_type
    outputprefix = args.output_prefix
    outputdirectory = args.output_directory

    if not os.path.exists(outputdirectory):
        os.makedirs(outputdirectory)

    def convert_to_excel(inputfile, outfile):
        wb = openpyxl.Workbook()
        ws = wb.active
        with open(inputfile, 'r') as f:
            for row in csv.reader(f):
                ws.append(row)
        wb.save(outfile)

    def copyRange(startCol, startRow, endCol, endRow, sheet):
        rangeSelected = []
        # Loops through selected Rows
        for i in range(startRow, endRow + 1, 1):
            # Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol, endCol+1, 1):
                rowSelected.append(sheet.cell(row=i, column=j).value)
            # Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
        return rangeSelected

    def copyRange_columnmajor(startCol, startRow, endCol, endRow, sheet):
        rangeSelected = []
        # Loops through selected Rows
        for j in range(startCol, endCol+1, 1):
            rowSelected = []
            for i in range(startRow, endRow + 1, 1):
                # Appends the row to a RowSelected list
                rowSelected.append(sheet.cell(row=i, column=j).value)
            # Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
        return rangeSelected

    # Paste range
    # Paste data from copyRange into template sheet
    def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
        countRow = 0
        for i in range(startRow, endRow+1, 1):
            countCol = 0
            for j in range(startCol, endCol+1, 1):

                sheetReceiving.cell(
                    row=i, column=j).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1

    def build_result(inputfile, template, plate_type, outputprefix):
        """
        1- Read Input File
        2- Read Template
        4- Depending on the plate type
            4-1 IF 96
                Read CSV input
                Copy values into destination cells
                Copy title (barcode to box)
                output is a template AND a log
            4-2 IF 48
                Read CSV input
                Copy values into destination cells
                Copy title (barcode to box)
                output is a template AND a log
            4-3 IF 196
                Read CSV input
                Copy values into destination cells
                Copy title (barcode to box)
                output is a template 

        """
        convert_to_excel(inputfile, 'tmp.xlsx')
        wb_input = openpyxl.load_workbook('tmp.xlsx')
        input_sheet = wb_input["Sheet"]

        template = openpyxl.load_workbook(template)
        output_sheet = template["Sheet1"]

        if plate_type == 96:
            # copy the grid
            selectedRange = copyRange(2, 4, 13, 11, input_sheet)
            pastingRange = pasteRange(
                2, 6, 13, 13, output_sheet, selectedRange)

            # copy the ID
            selectedRange = copyRange(1, 2, 1, 2, input_sheet)
            pastingRange = pasteRange(
                1, 3, 1, 3, output_sheet, selectedRange)

            template.save(os.path.join(outputdirectory, outputprefix+".xlsx"))

            # To print the log, we build a dataframe
            alphas = list(string.ascii_uppercase[:8])
            # ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
            IDs = []
            for i in range(1, 13):
                for a in alphas:
                    IDs.append(a+str("%.2d" % i))
            selectedRange = copyRange_columnmajor(2, 4, 13, 11, input_sheet)
            a = pd.DataFrame(selectedRange).values.ravel()
            data_tuples = list(zip(IDs, a))
            df = pd.DataFrame(data_tuples)
            df.to_csv(os.path.join(outputdirectory, outputprefix+".log"),
                      index=False, header=False, sep="\t")

        elif plate_type == 48:
            selectedRange = copyRange(2, 4, 9, 9, input_sheet)
            pastingRange = pasteRange(
                2, 6, 9, 11, output_sheet, selectedRange)
            # copy the ID
            selectedRange = copyRange(1, 2, 1, 2, input_sheet)
            pastingRange = pasteRange(
                1, 3, 1, 3, output_sheet, selectedRange)

            template.save(os.path.join(outputdirectory, outputprefix+".xlsx"))

            # To print the log, we build a dataframe
            alphas = list(string.ascii_uppercase[:6])
            # ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
            IDs = []
            for i in range(1, 9):
                for a in alphas:
                    IDs.append(a+str("%.2d" % i))
            selectedRange = copyRange_columnmajor(2, 4, 9, 9, input_sheet)
            a = pd.DataFrame(selectedRange).values.ravel()
            data_tuples = list(zip(IDs, a))
            df = pd.DataFrame(data_tuples)
            df.to_csv(os.path.join(outputdirectory, outputprefix+".log"),
                      index=False, header=False, sep="\t")

        elif plate_type == 196:
            selectedRange = copyRange(2, 3, 8, 30, input_sheet)
            a = pd.DataFrame(selectedRange).values.ravel()
            pastingRange = pasteRange(
                2, 2, 2, 197, output_sheet, a.astype(object).reshape(a.size, 1))
            # copy the ID
            #selectedRange = copyRange(1, 2, 1, 2, input_sheet)
            # pastingRange = pasteRange(
            #    1, 3, 1, 3, output_sheet, selectedRange)

            template.save(os.path.join(outputdirectory, outputprefix+".xlsx"))
        else:
            print("Error in Plate Type, please use either 96, 48 or 196 plate types")

        os.remove("tmp.xlsx")

    build_result(inputfile, template, plate_type, outputprefix)
    # print(df)
